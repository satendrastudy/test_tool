# %%
import pandas as pd
import numpy as np
import openpyxl
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from babel.numbers import format_currency
from openpyxl.styles import Border, Side, Alignment
import re

def def_natch_reco():
 df_natch_recon =fr'C:\Users\utk78122\OneDrive - Utkarsh Small Finance bank limited\Tanya CTS_NATCH Recon\NACH\NACH Audit Data Jul 2025\NACH Audit Recon Sheet for Jul 2025.xlsx'
 return df_natch_recon
for i in range(1, 32):
    j = str(i).zfill(2)
    df_natch_recon=def_natch_reco()
    print('file found',df_natch_recon)
    wb = load_workbook(df_natch_recon)# Load the existing workbook
    sheet = wb['Reconciliation']  # Select the desired sheet

    a=2+(i-1)*26
    b=3+(i-1)*26
    c=4+(i-1)*26
    d=5+(i-1)*26
    e=6+(i-1)*26
    f=7+(i-1)*26
    g=8+(i-1)*26
    h=9+(i-1)*26
    k=10+(i-1)*26
    l=11+(i-1)*26
    m=12+(i-1)*26
    n=13+(i-1)*26
    o=14+(i-1)*26
    p=15+(i-1)*26
    q=16+(i-1)*26
    r=17+(i-1)*26
    s=18+(i-1)*26
    t=19+(i-1)*26
    u=20+(i-1)*26
    v=21+(i-1)*26
    w=22+(i-1)*26
    x=23+(i-1)*26
    y=24+(i-1)*26
    z=25+(i-1)*26
# Write data to specific cells
    sheet.merge_cells(f'H{a}:L{a}')
    sheet[f'H{a}'] ='Audit NACH Reconciliation'
    sheet[f'H{a}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'H{a}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'H{a}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet[f'M{a}']=f'{j}-July-2025'
    sheet[f'M{a}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'M{a}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'M{a}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet.merge_cells(f'H{b}:J{b}')
    sheet[f'H{b}']=f'As per NACH Pooling'
    sheet[f'H{b}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'H{b}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'H{b}'].fill = PatternFill(start_color="7aace7", end_color="7aace7", fill_type="solid")
    sheet.merge_cells(f'K{b}:M{b}')
    sheet[f'K{b}']=f'As per RBI'
    sheet[f'K{b}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'K{b}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'K{b}'].fill = PatternFill(start_color="7aace7", end_color="7aace7", fill_type="solid")
    wb.save(df_natch_recon)
data_frames_emi = []

for i in range(1, 32):  # Loop from 1 to 31
    j = str(i).zfill(2)
    k = str(int(j) - 1).zfill(2)
    print(j)
    base_path = rf'C:\Users\utk78122\OneDrive - Utkarsh Small Finance bank limited\Tanya CTS_NATCH Recon\NACH\NACH Audit Data Jul 2025\{j} Jul, 2025\\'
    for l in range(1, 2):
        m = str(l).zfill(2)
        
        patterns = [
            f'ACH*.txt']
        
        for pattern in patterns:
            full_pattern = base_path + pattern
            for file_path in glob.glob(full_pattern):
                try:
                    df = pd.read_csv(file_path, delimiter='|', header=None, nrows=1)
                    df['Date'] = f'{j}-Jul-2025 {m}'
                    df['Date1'] = f'{j}-Jul-2025'
                    data_frames_emi.append(df)
                    print(f"File read successfully: {file_path}")
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")
                    continue
emi_coll_df = pd.concat(data_frames_emi, ignore_index=True)
emi_coll_df = emi_coll_df.rename(columns={0: 'Desp'})
emi_coll_df_MGL = emi_coll_df[emi_coll_df['Desp'].str.contains('MGL SAKH',case=False)]
emi_coll_df = emi_coll_df[~emi_coll_df['Desp'].str.contains('MGL SAKH', case=False,na=False)]
values_before_date = []
for i in range(len(emi_coll_df)):
    sample_str = emi_coll_df['Desp'].iloc[i]
    match = re.search(r'(\d+)(?=2025)', sample_str)
    if match:
        result = re.sub(r'^9+0*', '', match.group(1))
        #result = result[:-2] if len(result) > 2 else ''
        values_before_date.append(result)
    else:
        values_before_date.append(None)
emi_coll_df['Value_Before_date'] = [str(val) + '2025' if val is not None else None for val in values_before_date]
emi_coll_df['amount'] = emi_coll_df['Value_Before_date'].str[:-8]  # All digits except the last 8
emi_coll_df['Date'] = emi_coll_df['Value_Before_date'].str[-8:]    # Last 8 digits
emi_coll_df_grouped = emi_coll_df.groupby('Date1', as_index=False)['amount'].apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
# Function to add decimal before the last two digits
def format_amount(value):
    value_str = str(int(value))  # Ensure it's an integer string
    if len(value_str) < 3:
        value_str = value_str.zfill(3)
    return float(value_str[:-2] + '.' + value_str[-2:])
# Apply formatting to the 'amount' column
emi_coll_df_grouped['decimal_amount'] = emi_coll_df_grouped['amount'].apply(format_amount)
for i in range(1, 32):
    j = str(i).zfill(2)
    df_natch_recon=def_natch_reco()
    print('file found',df_natch_recon)
    wb = load_workbook(df_natch_recon)# Load the existing workbook
    sheet = wb['Reconciliation']  # Select the desired sheet
    t=19+(i-1)*26    
    sheet[f'H{t}']='Total EMI Collection'
    sheet[f'I{t}'] = format_currency(emi_coll_df_grouped.loc[(emi_coll_df_grouped['Date1'] == f'{j}-Jul-2025') ,'decimal_amount'].sum().astype(str), 'INR', locale='en_IN')
    wb.save(df_natch_recon)
# ### **NACH POOL**

data_frames1 = []
for i in range(1, 32):  # Loop from 1 to 31
    j = str(i).zfill(2)
    aa=1+(i-1)*26
    path = glob.glob(rf'C:\Users\utk78122\OneDrive - Utkarsh Small Finance bank limited\Tanya CTS_NATCH Recon\NACH\NACH POOL July\*{j}*.xls')
    if not path:
      continue
    try:
        # Attempt to read the Excel file
        df = pd.read_excel(path[0], header=9,sheet_name='rglledbc_xls').iloc[0:, 1:] 
        data_frames1.append(df)
        print(f'File read successfully:{path}')  # Append the DataFrame to the list
    except FileNotFoundError: 
        print(f"File not found: {path}. {FileNotFoundError} Skipping...")   
read_file1 = pd.concat(data_frames1, ignore_index=True) 
df = read_file1.replace('', None).dropna(how='any')
df = df.rename(columns={'Unnamed: 6': 'Tran_Amount', 'Tran Amount': 'Currency'})
df.columns = df.columns.str.strip()
df1 = df['Particulars']
deli = [':']
for col, dil in zip([df1], deli):  # Wrap df1 in a list to iterate correctly
    new = col.str.split(dil, expand=True)  # Use col directly as a Series
    new.columns = [f'{col.name}_{i+1}' for i in range(new.shape[1])]  
    df=df.join(new)# Use col.name for column naming
df23 = df[~(df['Particulars_1'] == 'ACH Debits')]
debit_sum_by_date1 = df23.groupby(['Date', 'Particulars_1'])[['Tran_Amount','Debits','Credits']].sum()
debit_sum_by_date1=pd.DataFrame(debit_sum_by_date1).reset_index()
Ach_D_df = df[df['Particulars'].str.contains('ACH Debits', na=False)]
Ach_D_df = Ach_D_df [~Ach_D_df ['Particulars'].str.contains('ACH Debits Return', na=False)]
Ach_D_df1 = Ach_D_df[
	~Ach_D_df['Tran Ref. No'].str.match(r'^\d{1,3}/', na=False) &
	Ach_D_df['Particulars_1'].str.contains(r'ACH Debits', na=False) & ~Ach_D_df['Particulars_1'].str.contains(r'ACH Debits Return', na=False) 
].drop_duplicates(subset=['Particulars'], keep='first')
debit_sum_by_date = Ach_D_df1.groupby(['Date', 'Particulars_1'])[['Tran_Amount','Debits','Credits']].sum()
debit_sum_by_date=pd.DataFrame(debit_sum_by_date).reset_index()
MerNach_df = pd.concat([debit_sum_by_date, debit_sum_by_date1])
MerNach_df = MerNach_df.sort_values(by=['Date'])
MerNach_df_DBT = MerNach_df[MerNach_df ['Particulars_1'].str.contains('DBT', na=False)]
MerNach_df = MerNach_df[~MerNach_df ['Particulars_1'].str.contains('DBT', na=False)]
MerNach_df_DBT=MerNach_df_DBT[MerNach_df_DBT['Particulars_1'].str.contains('NACH EMI COLLECTION', case=False,na=False)]
emi_coll_df_grouped=pd.merge(MerNach_df_DBT,emi_coll_df_grouped[['Date1','decimal_amount']],left_on='Date',right_on='Date1',how='left')
MerNach_df_DBT['final_tran']=abs(MerNach_df_DBT['Tran_Amount']-MerNach_df_DBT['decimal_amount'])
MerNach_df['Tran_Amount'] = MerNach_df['Tran_Amount'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
MerNach_df['Debits'] = MerNach_df['Debits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
MerNach_df['Credits'] = MerNach_df['Credits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
for i in range(1, 32):
    j = str(i).zfill(2)
    df_natch_recon=def_natch_reco()
    print('file found',df_natch_recon)
    wb = load_workbook(df_natch_recon)# Load the existing workbook
    sheet = wb['Reconciliation']  # Select the desired sheet
    a=2+(i-1)*26
    b=3+(i-1)*26
    c=4+(i-1)*26
    d=5+(i-1)*26
    e=6+(i-1)*26
    f=7+(i-1)*26
    g=8+(i-1)*26
    h=9+(i-1)*26
    k=10+(i-1)*26
    l=11+(i-1)*26
    m=12+(i-1)*26
    n=13+(i-1)*26
    o=14+(i-1)*26
    p=15+(i-1)*26
    q=16+(i-1)*26
    r=17+(i-1)*26
    s=18+(i-1)*26
    t=19+(i-1)*26
    u=20+(i-1)*26
    v=21+(i-1)*26
    w=22+(i-1)*26
    x=23+(i-1)*26
    y=24+(i-1)*26
    z=25+(i-1)*26
# Write data to specific cells
    sheet.merge_cells(f'H{a}:L{a}')
    sheet[f'H{a}'] ='Audit NACH Reconciliation'
    sheet[f'H{a}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'H{a}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'H{a}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet[f'M{a}']=f'{j}-Jul-2025'
    sheet[f'M{a}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'M{a}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'M{a}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet.merge_cells(f'H{b}:J{b}')
    sheet[f'H{b}']=f'As per NACH Pooling'
    sheet[f'H{b}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'H{b}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'H{b}'].fill = PatternFill(start_color="7aace7", end_color="7aace7", fill_type="solid")
    sheet.merge_cells(f'K{b}:M{b}')
    sheet[f'K{b}']=f'As per RBI'
    sheet[f'K{b}'].font = openpyxl.styles.Font(bold=True, size=12)
    sheet[f'K{b}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'K{b}'].fill = PatternFill(start_color="7aace7", end_color="7aace7", fill_type="solid")
    sheet[f'H{f}']='Total ACH Debit Transaction'
    value = MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'ACH Debits'),'Tran_Amount'].sum()
    sheet[f'I{f}'] =format_currency(value, 'INR', locale='en_IN')
    sheet[f'H{h}']='Total ACH Debit Return Transaction'
    sheet[f'I{h}'] = format_currency(MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'ACH Debits Return'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'H{l}']='Total ACH Credit Transaction'
    sheet[f'I{l}'] = format_currency(MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'NACH Credits'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'H{n}']='Total ACH Credit Return Transaction'
    sheet[f'I{n}'] = format_currency(MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'NACH Credits Return'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'H{p}']='Total APBS Credit Transaction'
    sheet[f'I{p}'] = format_currency(MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'APBS Credits'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'H{r}']='Total APBS Credit Return Transaction'
    sheet[f'I{r}'] = format_currency(MerNach_df.loc[(MerNach_df['Date'] == f'{j}-Jul-2025') & (MerNach_df['Particulars_1'] == 'APBS Credits Return'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'H{v}']='Total EMI Collection Return'
    sheet[f'I{v}'] = format_currency(MerNach_df_DBT.loc[(MerNach_df_DBT['Date'] == f'{j}-Jul-2025') ,'final_tran'].sum().astype(str), 'INR', locale='en_IN')
    wb.save(df_natch_recon)

def get_day_suffix(day):
    if 11 <= day <= 13:
        return 'th'
    last_digit = day % 10
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(last_digit, 'th')
# Base directory
base_dir_gl = Path(r"C:\Users\utk78122\OneDrive - Utkarsh Small Finance bank limited\Tanya CTS_NATCH Recon\NACH\RTGS SUSPENSE july")
# List to store DataFrames
data_frames2 = []
# Loop through days of July
for i in range(1, 32):
    suffix = get_day_suffix(i)
    file_name = f"{i}{suffix} July.xls"
    file_path = base_dir_gl / file_name
    try:
        df3 = pd.read_excel(file_path, header=9, sheet_name='rglledbc_xls')
        df3 = df3.iloc[:, 1:]  # Remove first column
        data_frames2.append(df3)
        print(f"✅ File read successfully: {file_path}")
    except FileNotFoundError:
        print(f"❌ File not found: {file_path}. Skipping...")
if data_frames2:
    read_file3 = pd.concat(data_frames2, ignore_index=True)
    print("✅ All files concatenated successfully.")
else:
    print("⚠️ No files were read. Please check the file paths.")
read_file3.columns = read_file3.columns.str.strip()
read_file3 = read_file3.rename(columns={'Unnamed: 6': 'Tran_Amount', 'Tran Amount': 'Currency'})
read_file3 = read_file3[read_file3['Particulars'].str.contains('MNSB', case=False, na=False)]
read_file3 = read_file3[~read_file3['Particulars'].str.contains('MNSB SETTLEMENT', case=False, na=False)]
read_file3 = read_file3[read_file3['Particulars'].str.contains(r'ACHSTL|APBSTL', case=False, na=False)]
df1 = read_file3['Particulars']
deli = [' ']
for col, dil in zip([df1], deli):  # Wrap df1 in a list to iterate correctly
    new = col.str.split(dil, expand=True)  # Use col directly as a Series
    new.columns = [f'{col.name}_{i+1}' for i in range(new.shape[1])]  
    df=read_file3.join(new)# Use col.name for column naming
columns_to_keep = ['Particulars_4', 'Particulars_5', 'Particulars_6']
columns_to_drop = [col for col in df.columns if col.startswith('Particulars_') and col not in columns_to_keep]
df_new = df.drop(columns=columns_to_drop)
df_new['Particulars_new'] = df_new['Particulars_4'].astype(str) + ' ' + df_new['Particulars_5'].astype(str) + ' ' + df_new['Particulars_6'].astype(str)
df_new.drop(columns=['Particulars_4', 'Particulars_5', 'Particulars_6'], inplace=True)
df_new.columns = df_new.columns.str.strip()
df23a = df_new[~(df_new['Particulars_new'] == 'ACH Debits')]
debit_sum_by_date_rbi1 = df23a.groupby(['Date', 'Particulars_new'])[['Tran_Amount', 'Debits', 'Credits']].sum()
debit_sum_by_date_rbi1 = debit_sum_by_date_rbi1.reset_index()
Ach_D_df_rbi_3 = df23a[df23a['Particulars'].str.contains('ACH Debit Return', na=False)]
df123 = Ach_D_df_rbi_3['Tran Ref. No']
deli = ['/']
for col, dil in zip([df123], deli):  # Wrap df1 in a list to iterate correctly
    new = col.str.split(dil, expand=True)  # Use col directly as a Series
    new.columns = [f'{col.name}_{i+1}' for i in range(new.shape[1])]  
    Ach_D_df_rbi_3=Ach_D_df_rbi_3.join(new)# Use col.name for column naming
Ach_D_df_rbi_31 = Ach_D_df_rbi_3 [Ach_D_df_rbi_3 ['Tran Ref. No_2'].str.contains('1', na=False)]
Ach_D_df_rbi_32 = Ach_D_df_rbi_3 [Ach_D_df_rbi_3 ['Tran Ref. No_2'].str.contains('2', na=False)]
Ach_D_df_rbi_31 = Ach_D_df_rbi_31.groupby(['Date', 'Particulars_new'])[['Tran_Amount', 'Debits', 'Credits']].sum()
Ach_D_df_rbi_31 = Ach_D_df_rbi_31.reset_index()
Ach_D_df_rbi_32 = Ach_D_df_rbi_32.groupby(['Date', 'Particulars_new'])[['Tran_Amount', 'Debits', 'Credits']].sum()
Ach_D_df_rbi_32 = Ach_D_df_rbi_32.reset_index()
Ach_D_df_rbi = df_new[df_new['Particulars_new'].str.contains('ACH Debit Presentation', na=False)]
Ach_D_df_rbi = Ach_D_df_rbi [~Ach_D_df_rbi ['Particulars'].str.contains('ACH Debit Return', na=False)]
df123 = Ach_D_df_rbi['Tran Ref. No']
deli = ['/']
for col, dil in zip([df123], deli):  # Wrap df1 in a list to iterate correctly
    new = col.str.split(dil, expand=True)  # Use col directly as a Series
    new.columns = [f'{col.name}_{i+1}' for i in range(new.shape[1])]  
    Ach_D_df_rbi=Ach_D_df_rbi.join(new)# Use col.name for column naming
Ach_D_df_rbi_22 = Ach_D_df_rbi [Ach_D_df_rbi ['Tran Ref. No_2'].str.contains('2', na=False)]
Ach_D_df_rbi_21 = Ach_D_df_rbi [Ach_D_df_rbi ['Tran Ref. No_2'].str.contains('1', na=False)]
Ach_D_df_rbi= Ach_D_df_rbi [~Ach_D_df_rbi ['Tran Ref. No_2'].str.contains('2', na=False)]
Ach_D_df_rbi.drop(columns=['Tran Ref. No_1'], inplace=True)
Ach_D_df_rbi_21 = Ach_D_df_rbi_21.groupby(['Date', 'Particulars_new'])[['Tran_Amount','Debits','Credits']].sum()
Ach_D_df_rbi_22 = Ach_D_df_rbi_22.groupby(['Date', 'Particulars_new'])[['Tran_Amount','Debits','Credits']].sum()
Ach_D_df_rbi_21 = Ach_D_df_rbi_21.reset_index()
Ach_D_df_rbi_22 = Ach_D_df_rbi_22.reset_index()
Ach_D_df_rbi_21['Tran_Amount'] = Ach_D_df_rbi_21['Tran_Amount'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
Ach_D_df_rbi_21['Debits'] = Ach_D_df_rbi_21['Debits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
Ach_D_df_rbi_21['Credits'] = Ach_D_df_rbi_21['Credits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
debit_sum_by_date_rbi = Ach_D_df_rbi.groupby(['Date', 'Particulars_new'])[['Tran_Amount','Debits','Credits']].sum()
debit_sum_by_date_rbi=pd.DataFrame(debit_sum_by_date_rbi).reset_index()
MerRBI_df1 = pd.concat([debit_sum_by_date_rbi, debit_sum_by_date_rbi1])
MerRBI_df1['Tran_Amount'] = MerRBI_df1['Tran_Amount'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
MerRBI_df1['Debits'] = MerRBI_df1['Debits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
MerRBI_df1['Credits'] = MerRBI_df1['Credits'].apply(lambda x: round(float(str(x).split()[-1]), 2) if pd.notnull(x) else x)
MerRBI_df1 = MerRBI_df1.sort_values(by=['Date'])
MerRBI_df1['Cr-Dr']=abs(MerRBI_df1['Credits']-MerRBI_df1['Debits'])
for i in range(1, 32):
    j = str(i).zfill(2)
    df_natch_recon=def_natch_reco()
    print('file found',df_natch_recon)
    wb = load_workbook(df_natch_recon)# Load the existing workbook
    sheet = wb['Reconciliation']  # Select the desired sheet
    a=2+(i-1)*26
    b=3+(i-1)*26
    c=4+(i-1)*26
    d=5+(i-1)*26
    e=6+(i-1)*26
    f=7+(i-1)*26
    g=8+(i-1)*26
    h=9+(i-1)*26
    k=10+(i-1)*26
    l=11+(i-1)*26
    m=12+(i-1)*26
    n=13+(i-1)*26
    o=14+(i-1)*26
    p=15+(i-1)*26
    q=16+(i-1)*26
    r=17+(i-1)*26
    s=18+(i-1)*26
    t=19+(i-1)*26
    u=20+(i-1)*26
    v=21+(i-1)*26
    w=22+(i-1)*26
    x=23+(i-1)*26
    y=24+(i-1)*26
    z=25+(i-1)*26
#outflow    
    sheet[f'K{f}']='Total of NACH Debit'
    value = Ach_D_df_rbi_21.loc[(Ach_D_df_rbi_21['Date'] == f'{j}-Jul-2025') & (Ach_D_df_rbi_21['Particulars_new'] == 'ACH Debit Presentation'),'Tran_Amount'].sum()
    sheet[f'L{f}'] =format_currency(value, 'INR', locale='en_IN')
    sheet[f'K{g}']='Total of NACH Credit Return'
    sheet[f'L{g}'] = format_currency(MerRBI_df1.loc[(MerRBI_df1['Date'] == f'{j}-Jul-2025') & (MerRBI_df1['Particulars_new'] == 'ACH Credit Return'),'Tran_Amount'].sum().astype(str),'INR', locale='en_IN')
    sheet[f'K{h}']='Total of APBS Credit Return'
    sheet[f'L{h}'] = format_currency(MerRBI_df1.loc[(MerRBI_df1['Date'] == f'{j}-Jul-2025') & (MerRBI_df1['Particulars_new'] == 'APB Credit Return'),'Tran_Amount'].sum().astype(str), 'INR', locale='en_IN')
    sheet[f'K{k}']='Total of NACH Debit Return (Collection)'
    sheet[f'L{k}'] = format_currency(Ach_D_df_rbi_31.loc[(Ach_D_df_rbi_31['Date'] == f'{j}-Jul-2025') & (Ach_D_df_rbi_31['Particulars_new'] == 'ACH Debit Return'),'Tran_Amount'].sum().astype(str),'INR', locale='en_IN')
#Inflow
    sheet[f'K{q}']='Total of NACH Debit Return'
    sheet[f'L{q}'] = format_currency(Ach_D_df_rbi_32.loc[(Ach_D_df_rbi_32['Date'] == f'{j}-Jul-2025') & (Ach_D_df_rbi_32['Particulars_new'] == 'ACH Debit Return'),'Tran_Amount'].sum().astype(str),'INR', locale='en_IN')
    sheet[f'K{r}']='Total of NACH Credit'
    sheet[f'L{r}'] = format_currency(MerRBI_df1.loc[(MerRBI_df1['Date'] == f'{j}-Jul-2025') & (MerRBI_df1['Particulars_new'] == 'ACH Credit Presentatio'),'Cr-Dr'].sum().astype(str),'INR', locale='en_IN')
    sheet[f'K{s}']='Total of APBS Credit'
    sheet[f'L{s}'] = format_currency(MerRBI_df1.loc[(MerRBI_df1['Date'] == f'{j}-Jul-2025') & (MerRBI_df1['Particulars_new'] == 'APB Credit Presentatio'),'Cr-Dr'].sum().astype(str),'INR', locale='en_IN')
    sheet[f'K{t}']='Total of NACH Debit (Collection)'
    sheet[f'L{t}'] = format_currency(Ach_D_df_rbi_22.loc[(Ach_D_df_rbi_22['Date'] == f'{j}-Jul-2025') & (Ach_D_df_rbi_22['Particulars_new'] == 'ACH Debit Presentation'),'Tran_Amount'].sum().astype(str),'INR', locale='en_IN')
#border    
    thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin') 
     )
    center_alignment = Alignment(horizontal='center', vertical='center')
    center_alignment2 = Alignment(horizontal='left', vertical='center')
    center_alignment3 = Alignment(horizontal='right', vertical='center')
# Apply the border to specific rows and columns
    rows_to_border = list(range(a, z + 1))   # Specify the rows
    columns_to_border = ['H', 'I', 'J','K','L','M']  # Specify the columns
    for row in rows_to_border:
          for col in columns_to_border:
           cell = sheet[f"{col}{row}"]
           cell.border = thin_border
           cell.alignment = center_alignment
    
    rows_to_border1 = list(range(f, z + 1))   # Specify the rows
    columns_to_border1 = ['H', 'J','K','M']  # Specify the columns
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['K'].width = 30
    for row in rows_to_border1:
          for col in columns_to_border1:
           cell = sheet[f"{col}{row}"]
           
           cell.alignment = center_alignment2
    rows_to_border2 = list(range(f, t + 1))   # Specify the rows
    columns_to_border2 = ['I', 'L']  # Specify the columns
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['I'].width = 15

    for row in rows_to_border2:
          for col in columns_to_border2:
           cell = sheet[f"{col}{row}"]
           
           cell.alignment = center_alignment3
    wb.save(df_natch_recon)



